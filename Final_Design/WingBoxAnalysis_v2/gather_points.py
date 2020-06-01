from openpyxl import load_workbook
import numpy as np
import scipy.interpolate as sp_interpolate
def gatherpoints(mesh, debug):
    if debug:
        print('*************** DEBUG MODE IS ON ***************')
        # airfoil_points_x1 = np.linspace(0, 1.5, 1500)
        # airfoil_points_x2 = np.linspace(1.5, 0.9, 600)
        # airfoil_points_x3 = np.linspace(0.9, 1.5, 600)
        # airfoil_points_x4 = np.linspace(1.5, 0 ,1500)
        # airfoil_points_z1 = np.linspace(0, 0.8, 1500)
        # airfoil_points_z2 = np.linspace(0.8,0, 600)
        # airfoil_points_z3 = np.linspace(0, - 0.8, 600)
        # airfoil_points_z4 = np.linspace(- 0.8,0, 1500)
        # airfoil_points_x = np.hstack([airfoil_points_x1, airfoil_points_x2, airfoil_points_x3, airfoil_points_x4])
        # airfoil_points_z = np.hstack([airfoil_points_z1, airfoil_points_z2, airfoil_points_z3, airfoil_points_z4])

        airfoil_points_x1 = np.linspace(0, 0.5, mesh)
        airfoil_points_x2 = np.linspace(0.5, 1, mesh)
        airfoil_points_x3 = np.linspace(1, 0, mesh)
        airfoil_points_z1 = np.linspace(0, 0.8, mesh)
        airfoil_points_z2 = np.linspace(0.8, 0, mesh)
        airfoil_points_z3 = np.zeros(mesh)
        airfoil_points_x = np.hstack([airfoil_points_x1, airfoil_points_x2, airfoil_points_x3])
        airfoil_points_z = np.hstack([airfoil_points_z1, airfoil_points_z2, airfoil_points_z3])


    else:

        wb = load_workbook(filename=r'CAL4014L_Points.xlsx')
        sheet = wb.worksheets[0]
        row_count = sheet.max_row  # count number of rows
        airfoil_points_x = []
        airfoil_points_z = []
        for i in range(2, row_count + 1):
            x_col = 'A'
            z_col = 'B'
            row_number = str(i)
            point_x_temp = sheet[x_col + row_number].value
            point_z_temp = sheet[z_col + row_number].value
            airfoil_points_x.append(point_x_temp)
            airfoil_points_z.append(point_z_temp)
        del row_number, point_x_temp, point_z_temp

        row_count = row_count - 1
        airfoil_points_x = np.array([airfoil_points_x]).T[:, 0]
        airfoil_points_z = np.array([airfoil_points_z]).T[:, 0]

        ###manipulate such that airfoil LE is 0,0
        airfoil_points_x = airfoil_points_x - airfoil_points_x.min()
        airfoil_points_z = airfoil_points_z - airfoil_points_z[airfoil_points_x == 0]

        slicing_temp = np.where(airfoil_points_x == 0)[0][0] + 1

        airfoil_points_z_upper_int = sp_interpolate.interp1d(airfoil_points_x[:slicing_temp],
                                                             airfoil_points_z[:slicing_temp])
        airfoil_points_z_lower_int = sp_interpolate.interp1d(airfoil_points_x[slicing_temp - 1:],
                                                             airfoil_points_z[slicing_temp - 1:])
        airfoil_points_x_temp = np.linspace(airfoil_points_x.min(), airfoil_points_x.max(), mesh)
        airfoil_points_z_upper_temp = airfoil_points_z_upper_int(airfoil_points_x_temp)
        airfoil_points_z_lower_temp = airfoil_points_z_lower_int(np.flip(airfoil_points_x_temp))

        airfoil_points_x = np.hstack([airfoil_points_x_temp, np.flip(airfoil_points_x_temp)])
        airfoil_points_z = np.hstack([airfoil_points_z_upper_temp, airfoil_points_z_lower_temp])

    # plt.clf()
    # # plt.figure(figsize=(9, 5))
    # ax = plt.gca()
    # # sets the ratio to 5
    # ax.set_aspect(1)
    # plt.plot(airfoil_points_x, airfoil_points_z, label='Cross section', color='r')
    # # plt.plot(np.linspace(0,c_loc, 150), np.zeros(150))
    # # plt.scatter([x_sc], [z_sc], label = 'Shear center', color = 'g')
    # # plt.scatter([x_bar], [z_bar], label = 'Centroid', color = 'b', marker = '*')
    # plt.legend(loc='upper right', prop={'size': 6})
    # plt.show()
    return airfoil_points_x, airfoil_points_z