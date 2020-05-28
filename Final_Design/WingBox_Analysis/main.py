"""
Author: Marco Desiderio

This code defines the 3D wing geometry

... complete description ...
"""
print('=========== INITIALIZING STRUCTURAL ANALYSIS COMPUTATIONS ===========')
import Load_CS as CS
import numpy as np
import matplotlib.pyplot as plt
import xlrd
import xlwt
from openpyxl import load_workbook
import parameters as par
import functions as f
import timeit
import Loads
import matplotlib
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.collections import LineCollection
from matplotlib.colors import ListedColormap, BoundaryNorm
import seaborn as sns
sns.set()

debug = False
plotcs = False
plotshow = False
plotsavefig = False
plotting = True

start = timeit.default_timer()
print('Coordinate system origin is located at the leading edge of the airfoil')
print('X-axis is pointing towards the TE and Z-axis is pointing up')
print('The CS origin depends of the input data')
# print('loading workbook...')
### load points from excel
wb = load_workbook(filename = r'NACA35120_Points.xlsx')
sheet = wb.worksheets[0]
# print('creating mesh...')
#print(sheet['A2'].value) # D18
row_count = sheet.max_row                   #count number of rows
airfoil_points_x = []
airfoil_points_z = []
for i in range(2, row_count + 1):
    x_col = 'A'
    z_col = 'B'
    row_number = str(i)
    point_x_loc = sheet[x_col + row_number].value
    point_z_loc = sheet[z_col + row_number].value
    airfoil_points_x.append(point_x_loc)
    airfoil_points_z.append(point_z_loc)


c_1 = par.c_r
x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc, airfoil_points_x_loc, airfoil_points_z_loc = CS.compute_CS_props(c_1, airfoil_points_x, airfoil_points_z, debug, plotcs, plotshow, plotsavefig)

### create chords list
c_locs = np.linspace(par.c_t, par.c_r, par.N)
c_locs_list = np.ndarray.tolist(c_locs)
### initialise arrays
x_bar_arr = np.zeros(1)
z_bar_arr = np.zeros(1)
Ixx_arr = np.zeros(1)
Izz_arr = np.zeros(1)
Izx_arr = np.zeros(1)
x_sc_arr = np.zeros(1)
z_sc_arr = np.zeros(1)
airfoil_points_x_arr = np.zeros(airfoil_points_x_loc.shape[0])
airfoil_points_z_arr = np.zeros(airfoil_points_x_loc.shape[0])
for c_loc in c_locs_list:
    x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc, airfoil_points_x_loc, airfoil_points_z_loc = CS.compute_CS_props(c_loc,
                                                                                                              airfoil_points_x,
                                                                                                              airfoil_points_z,
                                                                                                              debug,
                                                                                                              plotcs, plotshow, plotsavefig)
    x_bar_arr = np.append(x_bar_arr, x_bar)
    z_bar_arr = np.append(z_bar_arr, z_bar)
    Ixx_arr = np.append(Ixx_arr, Ixx)
    Izz_arr = np.append(Izz_arr, Izz)
    Izx_arr = np.append(Izx_arr, Izx)
    x_sc_arr = np.append(x_sc_arr, x_sc)
    z_sc_arr = np.append(z_sc_arr, z_sc)
    airfoil_points_x_arr = np.vstack((airfoil_points_x_arr, airfoil_points_x_loc))
    airfoil_points_z_arr = np.vstack((airfoil_points_z_arr, airfoil_points_z_loc))

x_bar_arr = x_bar_arr[1:]
z_bar_arr = z_bar_arr[1:]
Ixx_arr = Ixx_arr[1:]
Izz_arr = Izz_arr[1:]
Izx_arr = Izx_arr[1:]
x_sc_arr = x_sc_arr[1:]
z_sc_arr = z_sc_arr[1:]
airfoil_points_x_arr = airfoil_points_x_arr[1:,:]
airfoil_points_z_arr = airfoil_points_z_arr[1:,:]
Mx = Loads.Mx_array
# print(Mx.shape, airfoil_points_x_arr)

Mz = Loads.Mz_array
# Mx = np.zeros(Mz.shape[0])
sigma_y_arr = np.zeros(airfoil_points_x_arr.shape[1])
for i in range(par.N):
    i = int(i)
    sigma_y = f.get_bending_stresses(Mx[i], Mz[i], Ixx_arr[i], Izz_arr[i], Izx_arr[i], airfoil_points_x_arr[i,:], airfoil_points_z_arr[i,:], x_bar_arr[i], z_bar_arr[i])
    # print(sigma_y.shape)
    sigma_y_arr = np.vstack((sigma_y_arr, sigma_y))
sigma_y_arr = sigma_y_arr[1:,:]
# print(sigma_y_arr)

y_locs_arr = np.zeros(airfoil_points_x_arr.shape[1])
y_locs_pos = np.linspace(par.b/2, 0, par.N)

for i in range(par.N):
    i = int(i)
    y_locs_temp = y_locs_pos[i] * np.ones(airfoil_points_x_arr.shape[1])
    y_locs_arr = np.vstack((y_locs_arr, y_locs_temp))
y_locs_arr = y_locs_arr[1:,:]

## PLOT ROOT CHORD STRESSES###

sigma_y_root = sigma_y_arr[-1,:]/10e5
print(sigma_y_root.max(), sigma_y_root.min(), Mx[0])
print(z_bar_arr[0])
airfoil_points_x_arr_root = airfoil_points_x_arr[-1,:]
airfoil_points_z_arr_root = airfoil_points_z_arr[-1,:]


print('=========== STRUCTURAL ANALYSIS COMPUTATIONS COMPLETED ===========')
###time program execution
stop = timeit.default_timer()
print('Runtime: ', round(stop - start,4)*1000, 'ms')
print('==================================================================')


x = airfoil_points_x_arr_root
y = airfoil_points_z_arr_root
# dydx = np.cos(0.5 * (x[:-1] + x[1:]))  # first derivative

# Create a set of line segments so that we can color them individually
# This creates the points as a N x 1 x 2 array so that we can stack points
# together easily to get the segments. The segments array for line collection
# needs to be (numlines) x (points per line) x 2 (for x and y)
points = np.array([x, y]).T.reshape(-1, 1, 2)
segments = np.concatenate([points[:-1], points[1:]], axis=1)

fig, axs = plt.subplots(1, 1)
ax = plt.gca()
# sets the ratio to 5
ax.set_aspect(1)
# Create a continuous norm to map from data points to colors
norm = plt.Normalize(sigma_y_root.min(), sigma_y_root.max())
lc = LineCollection(segments, cmap='RdYlBu', norm=norm)
# Set the values used for colormapping
lc.set_array(sigma_y_root)
lc.set_linewidth(2)
line = axs.add_collection(lc)
fig.colorbar(line, ax=axs, label = '$\sigma_{y}$ [$MPa$]')
plt.xlabel('x [$m$]')
plt.ylabel('z [$m$]')
plt.minorticks_on()
plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)

axs.set_xlim(x.min() - 0.05, x.max() + 0.05)
axs.set_ylim(y.min() - 0.05, y.max() + 0.05)
plt.show()








#### !!!!!!!!! THIS TAKES WAAY TO MUCH TIME ##########
# # domains
# # x = np.logspace(-1.,np.log10(5),50)n # [0.1, 5]
# # y = np.linspace(6,9,50)             # [6, 9]
# # z = np.linspace(-1,1,50)            # [-1, 1]
# x = airfoil_points_x_arr.flatten()
# y = y_locs_arr.flatten()
# z = airfoil_points_z_arr.flatten()
# # convert to 2d matrices
# print('1')
# Z = np.outer(z.T, z)        # 50x50
# X, Y = np.meshgrid(x, y)    # 50x50
# # print(Y)
# print('2')
# # fourth dimention - colormap
# # create colormap according to x-value (can use any 50x50 array)
# color_dimension = X # change to desired fourth dimension
# minn, maxx = color_dimension.min(), color_dimension.max()
# norm = matplotlib.colors.Normalize(minn, maxx)
# m = plt.cm.ScalarMappable(norm=norm, cmap='jet')
# m.set_array([])
# fcolors = m.to_rgba(color_dimension)
# print('3')
# # plot
# fig = plt.figure()
# ax = fig.gca(projection='3d')
# print('Plotting')
# ax.plot_surface(X,Y,Z, rstride=1, cstride=1, facecolors=fcolors, vmin=minn, vmax=maxx, shade=False)
# print('Setting labels')
# ax.set_xlabel('x')
# ax.set_ylabel('y')
# ax.set_zlabel('z')
# # fig.canvas.show()
# plt.show()
