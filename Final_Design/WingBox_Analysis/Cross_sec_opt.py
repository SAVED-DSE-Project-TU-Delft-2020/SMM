"""
Author: Marco Desiderio

... insert description ...

"""
import numpy as np
from openpyxl import load_workbook
import Load_CS as CS
import matplotlib.pyplot as plt
import seaborn as sns
import scipy.interpolate as sp_interpolate
sns.set()


wb = load_workbook(filename = r'CAL4014L_Points.xlsx')
sheet = wb.worksheets[0]
# print('creating mesh...')
#print(sheet['A2'].value) # D18
row_count = sheet.max_row                   #count number of rows
airfoil_points_x = []
airfoil_points_z = []
for i in range(2, row_count + 1):
    x_col = 'A'
    z_col = 'B'
    row_number = str(i)
    point_x_loc = sheet[x_col + row_number].value
    point_z_loc = sheet[z_col + row_number].value
    airfoil_points_x.append(point_x_loc)
    airfoil_points_z.append(point_z_loc)
del row_number, point_x_loc, point_z_loc

row_count = row_count - 1

debug = False
plotcs = False
plotshow = False
plotsavefig = False





c_loc = 0.760

n_iter = 20
main_spar_locs = np.linspace(0.01, 0.25, n_iter)
main_spar_locs = np.ndarray.tolist(main_spar_locs)
aft_spar_loc = 0.75
# airfoil_per = 1.543689436179347
spar_cap_l = 0.040
spar_cap_t = 0.001
spar_cap_A = spar_cap_l * spar_cap_t
aft_spar_A = spar_cap_A/2
Ixxs = np.empty(1)
locs = np.empty(1)
for loc in main_spar_locs:
    airfoil_points_x_arr = np.array([airfoil_points_x])
    airfoil_points_x_arr[airfoil_points_x_arr < loc] = loc
    airfoil_points_x_arr[airfoil_points_x_arr > aft_spar_loc] = aft_spar_loc
    airfoil_points_x_ite = np.ndarray.tolist(airfoil_points_x_arr)[0]
    spar_loc_zs = []
    spar_locs_index = 0
    airfoil_points_z_temp = airfoil_points_z.copy()
    for point_x in airfoil_points_x_ite:
        if point_x == loc:

            spar_loc_zs.append(airfoil_points_z_temp[spar_locs_index])
        spar_locs_index = spar_locs_index + 1

    c_loc_ite = c_loc * (1 - (1 - aft_spar_loc) - loc)
    x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc, A_enclosed, airfoil_points_x_loc, airfoil_points_z_loc, mesh_length = CS.compute_CS_props(c_loc_ite, airfoil_points_x_ite, airfoil_points_z, debug, plotcs, plotshow, plotsavefig, spar_cap_A, aft_spar_A)

    z_up = spar_loc_zs[0]
    z_low = spar_loc_zs[-1]
    # print(z_up, z_low)
    Ixx = (Ixx + spar_cap_A * ((z_up - z_bar)**2 + (z_low - z_bar)**2))*10e11
    Ixxs = np.append(Ixxs, Ixx)
    locs = np.append(locs, loc)

locs = locs[1:]
Ixxs = Ixxs[1:]
plt.figure(figsize=(8,5))
plt.xlabel('Main spar location as a % of chord [-]')
plt.ylabel('$I_{xx}$ [$mm^4$]')
plt.title('Variation of second moment of area with different main spar locations, aft spar = 0.75c')
plt.minorticks_on()
plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.1)
plt.plot(locs, Ixxs, label = 'Second moment of area')
# plt.plot(locs, 0.95 * Ixxs[0] * np.ones(locs.shape[0]))
plt.scatter([locs[np.where(Ixxs == np.max(Ixxs))]], [np.max(Ixxs)], marker='o', label = 'Design point', edgecolors='r', facecolor = 'none')
plt.legend()
# plt.savefig('C:\\Users\\marco\\OneDrive\\Documents\\TU Delft\\BSc\\Year 3\\DSE\\Detailed Design\\Plots\\optimal_sparloc.pdf',dpi=600)
plt.show()
