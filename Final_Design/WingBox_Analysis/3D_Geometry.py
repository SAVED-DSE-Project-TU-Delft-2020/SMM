"""
Author: Marco Desiderio

This code defines the 3D wing geometry

... complete description ...
"""
import Load_CS as CS
import numpy as np
import matplotlib.pyplot as plt
import xlrd
import xlwt
from openpyxl import load_workbook
import parameters as par
import functions as f

debug = True
plotcs = True

print('=========== INITIALIZING STRUCTURAL ANALYSIS COMPUTATIONS ===========')
print('Coordinate system origin is located at the leading edge of the airfoil')
print('X-axis is pointing towards the TE and Z-axis is pointing up')
print('The CS origin depends of the input data')
# print('loading workbook...')
### load points from excel
wb = load_workbook(filename = r'NACA0012_Points.xlsx')
sheet = wb.worksheets[0]
# print('creating mesh...')
#print(sheet['A2'].value) # D18
row_count = sheet.max_row                   #count number of rows
airfoil_points_x = []
airfoil_points_z = []
for i in range(2, row_count + 1):
    x_col = 'A'
    z_col = 'B'
    row_number = str(i)
    point_x_loc = sheet[x_col + row_number].value
    point_z_loc = sheet[z_col + row_number].value
    airfoil_points_x.append(point_x_loc)
    airfoil_points_z.append(point_z_loc)


c_1 = 0.55

x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc = CS.compute_CS_props(c_1, airfoil_points_x, airfoil_points_z, debug, plotcs)




