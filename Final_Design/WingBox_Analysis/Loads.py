"""
Author: Marco Desiderio
This file computes the loads acting over the wing. Lift distribution is an input from the Aerodynamics department and
the weight distribution is estimated in a similar fashion as in the midterm report.

"""
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import scipy.integrate as sp_integrate
import scipy.interpolate as sp_interpolate
import parameters as par
import functions as f
import seaborn as sns
sns.set()

debug = False

print('=========== GENERATING LOAD CASES ===========')
####################### GET LIFT DISTRIBUTION FROM DATA ############################
wb = load_workbook(filename = r'testdata.xlsx')
sheet = wb.worksheets[0]
row_count = sheet.max_row                   #count number of rows
y_locs = []
dL_dy = []
for i in range(2, row_count + 1):
    locs_col = 'B'
    lift_col = 'C'
    row_number = str(i)
    point_y_locs = sheet[locs_col + row_number].value
    point_dL_dy_loc = sheet[lift_col + row_number].value
    y_locs.append(point_y_locs)
    dL_dy.append(point_dL_dy_loc)
###### MODIFY THIS WHEN WE GET NEW DATA ######
y_locs = np.array([y_locs])[0][:35]
dL_dy = np.array([dL_dy])[0][:35]

y_locs = np.hstack([[0], y_locs])
dL_dy = np.hstack([[dL_dy[0]], dL_dy])
dL_dy_new = sp_interpolate.interp1d(y_locs, dL_dy)      ### INTERPOLATE SO WE CAN GET THE DATA WHERE WE WANT

################## COMPUTE DOWNWARDS LOADS DUE TO STRCTURAL WEIGHT #################################

start = 0
stop = par.b/2
step = 0.001
num = f.computenum(start, stop, step)
y = np.linspace(start, stop, num)
Ax = par.c_r - y * par.c_r / par.h

M_CENTER = par.M_AI + par.M_BAT + par.M_PAY + par.M_PARACH + par.M_FINS

M_wing = par.MTOM - M_CENTER
M_wing_2 = M_wing / 2

wing_load = Ax * M_wing_2 / f.findarea(par.PAY_WIDTH/2, par.b/2, par.c_r, par.h) * 9.81
slicing = int(0.5 * par.PAY_WIDTH/step)
wing_load = wing_load[slicing:]
center_load = M_CENTER / 2 / (par.c_r * par.PAY_WIDTH / 2) * np.ones(slicing) * 9.81
w_mass_tot = np.hstack([center_load, wing_load])
mass = np.trapz(w_mass_tot, y) * 2 / 9.81
w_mass_tot = w_mass_tot * par.MTOM / mass

w_final = dL_dy_new(y) - w_mass_tot

############### TORSIONAL LOADS #############



if debug:
    plt.xlabel('y [$m$]')
    plt.ylabel('w [$N/m$]')
    plt.title('Spanwise load distribution (half-wing)')
    plt.minorticks_on()
    plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.1)
    plt.plot(y, w_mass_tot, label = 'Weight distribution')
    plt.plot(y, dL_dy_new(y), label = 'Lift distribution')
    plt.plot(y, w_final, label = 'Total distribution')
    plt.legend(loc = 'upper right')
    plt.show()