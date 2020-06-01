"""
Author: Marco Desiderio

... complete description ...
"""
print('=========== INITIALIZING STRUCTURAL ANALYSIS COMPUTATIONS ===========')
import Load_CS as CS
import numpy as np
import matplotlib.pyplot as plt
import xlrd
import xlwt
from openpyxl import load_workbook
import parameters as par
import functions as f
import timeit
import Loads
import matplotlib
from mpl_toolkits.mplot3d import Axes3D
from matplotlib.collections import LineCollection
from matplotlib.colors import ListedColormap, BoundaryNorm
import seaborn as sns
import Interal_Loads
import gc
gc.collect()
sns.set()

debug = False
plotcs = False
plotshow = False
plotsavefig = False
plotting = True

start = timeit.default_timer()
print('Coordinate system origin is located at the leading edge of the airfoil')
print('X-axis is pointing towards the TE and Z-axis is pointing up')
print('The CS origin depends of the input data')
# print('loading workbook...')
### load points from excel
wb = load_workbook(filename = r'CAL4014L_Points.xlsx')
sheet = wb.worksheets[0]
# print('creating mesh...')
#print(sheet['A2'].value) # D18
row_count = sheet.max_row                   #count number of rows
airfoil_points_x = []
airfoil_points_z = []
for i in range(2, row_count + 1):
    x_col = 'A'
    z_col = 'B'
    row_number = str(i)
    point_x_loc = sheet[x_col + row_number].value
    point_z_loc = sheet[z_col + row_number].value
    airfoil_points_x.append(point_x_loc)
    airfoil_points_z.append(point_z_loc)
del row_number, point_x_loc, point_z_loc

row_count = row_count - 1
if debug:
    c_1 = par.c_r
    x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc, A_enclosed, airfoil_points_x_loc, airfoil_points_z_loc, mesh_length = CS.compute_CS_props(c_1, airfoil_points_x, airfoil_points_z, debug, plotcs, plotshow, plotsavefig)
    row_count = airfoil_points_x_loc.shape[0]

main_spar_loc = par.main_spar_loc
aft_spar_loc = par.aft_spar_loc
mainspar_cap_l = par.mainspar_cap_l
mainspar_cap_t = par.mainspar_cap_t
aftspar_cap_l = par.aftspar_cap_l
aftspar_cap_t = par.aftspar_cap_t
mainspar_cap_A = mainspar_cap_l * mainspar_cap_t
aftspar_cap_A = aftspar_cap_l * aftspar_cap_t

### redefine shape according to spar locs
airfoil_points_x_arr = np.array([airfoil_points_x])
airfoil_points_x_arr[airfoil_points_x_arr < main_spar_loc] = main_spar_loc
airfoil_points_x_arr[airfoil_points_x_arr > aft_spar_loc] = aft_spar_loc
airfoil_points_x = np.ndarray.tolist(airfoil_points_x_arr)[0]


### create chords list
c_locs = np.linspace(par.c_t, par.c_r, par.N)
c_locs_list = np.ndarray.tolist(c_locs)
### initialise arrays
x_bar_arr = np.zeros(1)
z_bar_arr = np.zeros(1)
Ixx_arr = np.zeros(1)
Izz_arr = np.zeros(1)
Izx_arr = np.zeros(1)
x_sc_arr = np.zeros(1)
z_sc_arr = np.zeros(1)
A_m_arr = np.zeros(1)
airfoil_points_x_arr = np.zeros(row_count)
airfoil_points_z_arr = np.zeros(row_count)
mesh_length_arr = np.zeros(row_count)
for c_loc in c_locs_list:
    c_loc_ite = round(c_loc * (1 - (1 - aft_spar_loc) - main_spar_loc),5)
    x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc, A_enclosed, airfoil_points_x_loc, airfoil_points_z_loc, mesh_length = CS.compute_CS_props(c_loc_ite,
                                                                                                              airfoil_points_x,
                                                                                                              airfoil_points_z,
                                                                                                              debug,
                                                                                                              plotcs, plotshow, plotsavefig, mainspar_cap_A, aftspar_cap_A)

    x_bar_arr = np.append(x_bar_arr, x_bar)
    z_bar_arr = np.append(z_bar_arr, z_bar)

    main_spar_upperz = np.max(airfoil_points_z_loc[airfoil_points_x_loc == 0])
    main_spar_lowerz = np.min(airfoil_points_z_loc[airfoil_points_x_loc == 0])
    aft_spar_upperz = np.max(airfoil_points_z_loc[airfoil_points_x_loc == np.max(airfoil_points_x_loc)])
    aft_spar_lowerz = np.min(airfoil_points_z_loc[airfoil_points_x_loc == np.max(airfoil_points_x_loc)])

    Ixx = Ixx + mainspar_cap_A * ((main_spar_upperz - z_bar)**2 + (main_spar_lowerz - z_bar)**2) + aftspar_cap_A * ((aft_spar_upperz - z_bar)**2 + (aft_spar_lowerz - z_bar)**2)
    Izz = Izz + 2 * mainspar_cap_A * (z_bar**2) + 2 * aftspar_cap_A * (np.max(airfoil_points_x_loc) - x_bar)**2
    Ixx_arr = np.append(Ixx_arr, Ixx)
    Izz_arr = np.append(Izz_arr, Izz)
    Izx_arr = np.append(Izx_arr, Izx)
    x_sc_arr = np.append(x_sc_arr, x_sc)
    z_sc_arr = np.append(z_sc_arr, z_sc)
    A_m_arr = np.append(A_m_arr, A_enclosed)
    airfoil_points_x_arr = np.vstack((airfoil_points_x_arr, airfoil_points_x_loc))
    airfoil_points_z_arr = np.vstack((airfoil_points_z_arr, airfoil_points_z_loc))
    mesh_length_arr = np.vstack((mesh_length_arr, mesh_length))
## delete items to clear some memory and avoid mistakes
del x_bar, z_bar, Ixx, Izz, Izx, x_sc, z_sc, airfoil_points_x_loc, airfoil_points_z_loc, A_enclosed, mesh_length



x_bar_arr = x_bar_arr[1:]
z_bar_arr = z_bar_arr[1:]
Ixx_arr = Ixx_arr[1:]
Izz_arr = Izz_arr[1:]
Izx_arr = Izx_arr[1:]
x_sc_arr = x_sc_arr[1:]
z_sc_arr = z_sc_arr[1:]
A_m_arr = A_m_arr[1:]
airfoil_points_x_arr = airfoil_points_x_arr[1:,:]
airfoil_points_z_arr = airfoil_points_z_arr[1:,:]
mesh_length_arr = mesh_length_arr[1:,:]



#### GETHER INTERNAL LOADS ####
Mx = Interal_Loads.Mx_array
Mz = Interal_Loads.Mz_array
Sx = Interal_Loads.Sx_array
Sz = Interal_Loads.Sz_array
# insert Ty due to pitching moment

#### COMPUTE TORSION DUE TO AC =/ SC ####
x_ac_arr = 0.25 * c_locs
z_ac_arr = (airfoil_points_z_arr[:, 40] + airfoil_points_z_arr[:, -41]) / 2
Sx_arm = np.abs(z_ac_arr - z_sc_arr)
Sz_arm = np.abs(x_ac_arr - x_sc_arr)

Ty_Sx = Sx * Sx_arm
Ty_Sz = Sz * Sz_arm

#### IMPORT TORSION DUE TO PITCHING MOMENT ####
### this is missing ###

#######################################################################
Ty = np.abs(Ty_Sx + Ty_Sz)  #insert Ty due to pitching moment


sigma_yy_arr = np.zeros(row_count)
q_T_arr = np.zeros(1)
for i in range(par.N):
    i = int(i)
    sigma_yy = f.get_bending_stresses(Mx[i], Mz[i], Ixx_arr[i], Izz_arr[i], Izx_arr[i], airfoil_points_x_arr[i,:], airfoil_points_z_arr[i,:], x_bar_arr[i], z_bar_arr[i])
    q_T = Ty[i]/(2 * A_m_arr[i])
    # print(sigma_y.shape)
    sigma_yy_arr = np.vstack((sigma_yy_arr, sigma_yy))
    q_T_arr = np.vstack((q_T_arr, q_T))

del sigma_yy, q_T
sigma_yy_arr = sigma_yy_arr[1:,:]
q_T_arr = q_T_arr[1:]

##### SHEAR FLOW IS POSITIVE COUNTERCLOCKWISE - IF WE LOOK AT THE AIRFOIL WHILE KEEPING THE LE ON THE LEFT
#### compute shear flows due to shear forces
###find_qb(Sx, Sz, Ixx, Izz, Izx, line_coordinates, t, x, z)
###find_q0(qb, line_coordinates, skin_perimeter):
qb_arr = np.zeros(row_count)
for i in range(par.N):
    i = int(i)
    x = airfoil_points_x_arr[i,:] - x_bar_arr[i]
    z = airfoil_points_z_arr[i,:] - z_bar_arr[i]
    line_coordinates = np.cumsum(mesh_length_arr[i,:])
    shearz = 0
    qb_temp = f.find_qb(Sx[i], shearz, Ixx_arr[i], Izz_arr[i], Izx_arr[i], line_coordinates, par.t_sk, x, z )
    q0_temp = f.find_q0(qb_temp, line_coordinates, np.sum(mesh_length_arr[i,:]))
    q_temp = qb_temp + q0_temp
    qb_arr = np.vstack((qb_arr, q_temp))
del q_temp, q0_temp, qb_temp

qb_arr = qb_arr[1:,:]

q_tot_arr = np.zeros(row_count)
for i in range(par.N):
    i = int(i)
    q_tot_temp = qb_arr[i,:] + q_T_arr[i]
    q_tot_arr = np.vstack((q_tot_arr, q_tot_temp))
del q_tot_temp

q_tot_arr = q_tot_arr[1:,:]

tau_xy_arr = q_tot_arr / 10e5 / par.t_sk

y_locs_arr = np.zeros(row_count)
y_locs_pos = np.linspace(par.b/2, par.PAY_WIDTH/2, par.N)

for i in range(par.N):
    i = int(i)
    y_locs_temp = y_locs_pos[i] * np.ones(row_count)
    y_locs_arr = np.vstack((y_locs_arr, y_locs_temp))

del y_locs_temp

y_locs_arr = y_locs_arr[1:,:]


## PLOT STRESSES###
location = -1
sigma_yy_loc = sigma_yy_arr[location,:]/10e5
# print(sigma_y_root.max(), sigma_y_root.min(), Mx[0])
# print(z_bar_arr[0])
airfoil_points_x_arr_root = airfoil_points_x_arr[location,:]
airfoil_points_z_arr_root = airfoil_points_z_arr[location,:]
y_loc = y_locs_arr[location,0]


#### COMPUTE VON MISES STRESSES ####

### for now we define some stresses to be zero, we will change this once we have values
sigma_xx_loc = sigma_yy_loc * 0
sigma_zz_loc = sigma_yy_loc * 0
tau_xy_loc = tau_xy_arr[location]
tau_yz_loc = sigma_yy_loc * 0
tau_zx_loc = sigma_yy_loc * 0
sigma_vm = np.sqrt(((sigma_yy_loc + sigma_xx_loc)**2 + (sigma_yy_loc - sigma_zz_loc)**2 + (sigma_zz_loc - sigma_xx_loc)**2) / 2 + 3 * (tau_xy_loc**2 + tau_yz_loc**2 + tau_zx_loc**2))
print('=========== STRUCTURAL ANALYSIS COMPUTATIONS COMPLETED ===========')
###time program execution
stop = timeit.default_timer()
print('Runtime: ', round(stop - start,3)*1000, 'ms')
print('==================================================================')


x = airfoil_points_x_arr_root
y = airfoil_points_z_arr_root
# dydx = np.cos(0.5 * (x[:-1] + x[1:]))  # first derivative

# Create a set of line segments so that we can color them individually
# This creates the points as a N x 1 x 2 array so that we can stack points
# together easily to get the segments. The segments array for line collection
# needs to be (numlines) x (points per line) x 2 (for x and y)
points = np.array([x, y]).T.reshape(-1, 1, 2)
segments = np.concatenate([points[:-1], points[1:]], axis=1)
print('Plotting ...')
fig, axs = plt.subplots(1, 1)
ax = plt.gca()
# sets the ratio
ax.set_aspect(1)
# Create a continuous norm to map from data points to colors
# norm = plt.Normalize(sigma_yy_loc.min(), sigma_yy_loc.max())
norm = plt.Normalize(sigma_vm.min(), sigma_vm.max())
# norm = plt.Normalize(tau_xy_loc.min(), tau_xy_loc.max())
# lc = LineCollection(segments, cmap='RdYlBu', norm=norm)
lc = LineCollection(segments, cmap='RdYlBu_r', norm=norm)
# Set the values used for colormapping
# lc.set_array(sigma_yy_loc)
lc.set_array(sigma_vm)
# lc.set_array(tau_xy_loc)
lc.set_linewidth(3)
line = axs.add_collection(lc)
fig.colorbar(line, ax=axs, label = '$\sigma_{vm}$ [$MPa$]', orientation = 'horizontal', ticks = np.round(np.linspace(sigma_vm.min(), sigma_vm.max(), 15), 2))
plt.xlabel('x [$m$]')
plt.ylabel('z [$m$]')
plt.title('$\sigma_{yy}$ distribution along the airfoil, y = ' + str(y_loc) + 'm')
plt.minorticks_on()
plt.grid(b=True, which='minor', color='#999999', linestyle='-', alpha=0.2)
plt.scatter(x_sc_arr[location], z_sc_arr[location], marker='*', label = 'Shear center')
plt.scatter(x_ac_arr[location], z_ac_arr[location], label = 'Aerodynamic center')
plt.scatter(x_bar_arr[location], z_bar_arr[location], label = 'Centroid', marker='1')
axs.set_xlim(x.min() - 0.05, x.max() + 0.05)
axs.set_ylim(y.min() - 0.05, y.max() + 0.05)
plt.legend()
# plt.savefig('test.pdf')
plt.show()








#### !!!!!!!!! THIS TAKES WAAY TO MUCH TIME ##########
# # domains
# # x = np.logspace(-1.,np.log10(5),50)n # [0.1, 5]
# # y = np.linspace(6,9,50)             # [6, 9]
# # z = np.linspace(-1,1,50)            # [-1, 1]
# x = airfoil_points_x_arr.flatten()
# y = y_locs_arr.flatten()
# z = airfoil_points_z_arr.flatten()
# # convert to 2d matrices
# print('1')
# Z = np.outer(z.T, z)        # 50x50
# X, Y = np.meshgrid(x, y)    # 50x50
# # print(Y)
# print('2')
# # fourth dimention - colormap
# # create colormap according to x-value (can use any 50x50 array)
# color_dimension = X # change to desired fourth dimension
# minn, maxx = color_dimension.min(), color_dimension.max()
# norm = matplotlib.colors.Normalize(minn, maxx)
# m = plt.cm.ScalarMappable(norm=norm, cmap='jet')
# m.set_array([])
# fcolors = m.to_rgba(color_dimension)
# print('3')
# # plot
# fig = plt.figure()
# ax = fig.gca(projection='3d')
# print('Plotting')
# ax.plot_surface(X,Y,Z, rstride=1, cstride=1, facecolors=fcolors, vmin=minn, vmax=maxx, shade=False)
# print('Setting labels')
# ax.set_xlabel('x')
# ax.set_ylabel('y')
# ax.set_zlabel('z')
# # fig.canvas.show()
# plt.show()
